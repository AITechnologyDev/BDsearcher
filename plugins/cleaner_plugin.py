import os
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

def clean_file(file_path):
    ext = os.path.splitext(file_path)[1].lower()
    
    if ext == ".xlsx":
        return clean_excel(file_path)
    
    if ext not in (".txt", ".md", ".csv"):
        return False, "Unsupported file type"
    
    return clean_text_file(file_path)

def clean_text_file(file_path):
    try:
        with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
            lines = f.readlines()
            
        cleaned = set()
        cleaned_lines = []
        for line in lines:
            clean_line = line.strip()
            if clean_line and clean_line.lower() not in ("error", "null", "none", "undefined") and clean_line not in cleaned:
                cleaned.add(clean_line)
                cleaned_lines.append(line)
                
        with open(file_path, 'w', encoding='utf-8') as f:
            f.writelines(cleaned_lines)
            
        return True, f"Cleaned {file_path}: {len(lines) - len(cleaned_lines)} duplicates/garbage removed"
    
    except Exception as e:
        return False, f"Cleaning failed: {e}"

def clean_excel(file_path):
    try:
        wb = load_workbook(file_path)
        removed_count = 0
        cleaned_data = []

        for sheet in wb:
            unique_rows = []
            seen_rows = set()
            
            for row in sheet.iter_rows(values_only=True):
                # Проверка на пустые и мусорные строки
                if all(cell is None or str(cell).strip() == '' for cell in row):
                    removed_count += 1
                    continue
                    
                # Проверка на запрещенные значения
                if any(str(cell).lower() in ("error", "null", "none", "undefined") for cell in row if cell):
                    removed_count += 1
                    continue
                
                # Удаление дубликатов
                row_tuple = tuple(str(cell).strip().lower() for cell in row)
                if row_tuple not in seen_rows:
                    seen_rows.add(row_tuple)
                    unique_rows.append(row)

            # Очищаем лист и записываем уникальные данные
            sheet.delete_rows(1, sheet.max_row)
            for row in unique_rows:
                sheet.append(row)
                
            cleaned_data.append((sheet.title, len(unique_rows)))

        wb.save(file_path)
        summary = ", ".join([f"{name}: {rows} rows" for name, rows in cleaned_data])
        return True, f"Cleaned Excel: {summary} | Removed: {removed_count} rows"
        
    except Exception as e:
        return False, f"Excel cleaning failed: {str(e)}"

def run_cleaner():
    import glob
    from rich.console import Console
    console = Console()
    
    path = input("Enter directory to clean (default: .): ").strip() or "."
    found = 0
    
    # Расширенные паттерны поиска
    patterns = ("*.txt", "*.md", "*.csv", "*.xlsx")
    
    for pattern in patterns:
        for file in glob.glob(os.path.join(path, "**", pattern), recursive=True):
            ok, msg = clean_file(file)
            if ok:
                found += 1
                console.print(f"[green]{msg}[/]")
            else:
                console.print(f"[yellow]{msg}[/]")
    
    if found == 0:
        console.print("[yellow]No files cleaned[/]")

def register(registry):
    registry["cleaner"] = {
        "handler": run_cleaner,
        "name": "Enhanced Cleaner",
        "description": "Removes duplicates and garbage from txt/md/csv/xlsx files"
    }