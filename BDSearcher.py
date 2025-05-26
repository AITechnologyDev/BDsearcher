import os
import sys
import os
import sys
import json
import csv
import sqlite3
from datetime import datetime
from concurrent.futures import ThreadPoolExecutor
from rich.console import Console
from rich.panel import Panel
from rich.text import Text
from rich.table import Table
from openpyxl import load_workbook
from typing import List, Dict, Tuple
import requests
import subprocess
import re

console = Console()
highlight_style = "bold yellow on dark_red"
SUPPORTED_EXTENSIONS = {'.txt', '.sql', '.xlsx', '.csv', '.json','.md','.pdf','db', '.zip', '.rar'}
MAX_WORKERS = 4 if hasattr(sys, 'getandroidapilevel') else os.cpu_count()
REPORTS_DIR = "search_reports"
PLUGINS_DIR = "plugins"

# После импортов
if not os.path.exists(PLUGINS_DIR):
    os.makedirs(PLUGINS_DIR)
    
file_handlers = {}  # ключ: расширение, значение: функция

# API ключи и модели (заполните своими данными)
API_KEYS = {
    'huggingface': 'hf_xxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx',
    'google': 'AIxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxxx'
}

MODEL_NAMES = {
    'huggingface': 'deepseek-ai/DeepSeek-R1',
    'google': 'gemini-2.0-flash'
}

SERVER_TYPE = ()

# Путь к файлу истории
HISTORY_FILE = 'chat_history.json'

def load_history():
    if os.path.exists(HISTORY_FILE):
        with open(HISTORY_FILE, 'r', encoding='utf-8') as f:
            return json.load(f)
    return []

def save_history(messages):
    with open(HISTORY_FILE, 'w', encoding='utf-8') as f:
        json.dump(messages, f, ensure_ascii=False, indent=2)

def add_message(messages, role, content):
    messages.append({"role": role, "content": content})
    save_history(messages)



# Определение мобильной платформы
IS_MOBILE = hasattr(sys, 'getandroidapilevel')
MOBILE_WIDTH = 40 if IS_MOBILE else None

# Загрузка плагинов
def load_plugins():
    import importlib.util
    if not os.path.exists(PLUGINS_DIR):
        os.makedirs(PLUGINS_DIR)
    for file in os.listdir(PLUGINS_DIR):
        if file.endswith('.py'):
            path = os.path.join(PLUGINS_DIR, file)
            spec = importlib.util.spec_from_file_location(file[:-3], path)
            module = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(module)
            if hasattr(module, "register"):
                module.register(file_handlers)
                
def show_plugins_info():
    """Display plugin information, then wait for user to press Enter"""
    clear_screen()
    if not file_handlers:
        console.print("[yellow]No plugins loaded.[/]")
    else:
        table = Table(title="Plugins Information")
        table.add_column("Extension", style="cyan")
        table.add_column("Name", style="green")
        table.add_column("Description", style="magenta")
        for ext, info in file_handlers.items():
            name = info.get('name', '-')
            description = info.get('description', '-')
            table.add_row(ext, name, description)
        console.print(table)
    console.print("\nPress [bold green]Enter[/] to return to the main menu...")
    input()



def ask_ai(prompt, server_type=SERVER_TYPE):
    messages = load_history()
    add_message(messages, "user", prompt)

    if server_type == 'huggingface':
        API_URL = "https://router.huggingface.co/nebius/v1/chat/completions"
        headers = {
            "Authorization": f"Bearer {API_KEYS['huggingface']}",
            "Content-Type": "application/json"
        }
        payload = {
            "messages": [{"role": "user", "content": prompt}],
            "model": MODEL_NAMES['huggingface'], 
            "max_tokens": 4096,
            "temperature": 0.5
        }

        try:
            response = requests.post(API_URL, headers=headers, json=payload, timeout=300)
            
            if response.status_code == 200:
                answer = response.json()["choices"][0]["message"]["content"]
            else:
                error = response.json().get("error", {}).get("message", "Unknown error")
                answer = f"API Error [{response.status_code}]: {error}"
                
        except requests.exceptions.RequestException as e:
            answer = f"🌐 Network error: {str(e)}"
        except (KeyError, IndexError) as e:
            answer = f"🔍 Response format error: {str(e)}"
        except Exception as e:
            answer = f"💥 Critical: {str(e)}"

    elif server_type == 'google':
        try:
            url = f"https://generativelanguage.googleapis.com/v1beta/models/{MODEL_NAMES['google']}:generateContent"
            headers = {
                "Content-Type": "application/json",
                "x-goog-api-key": API_KEYS['google'].strip()
            }
            payload = {
                "contents": [{
                    "parts": [{"text": prompt}]
                }],
                "generationConfig": {
                    "temperature": 0.7,
                    "maxOutputTokens": 4096
                }
            }
            
            response = requests.post(url, headers=headers, json=payload, timeout=300)
            
            if response.status_code == 200:
                answer = response.json()["candidates"][0]["content"]["parts"][0]["text"]
            else:
                error_data = response.json().get("error", {})
                answer = f"🚨 Error Gemini [{response.status_code}]: {error_data.get('message', 'Unknown error')}"

        except requests.exceptions.RequestException as e:
            answer = f"🌐 Network error: {str(e)}"
        except (KeyError, IndexError) as e:
            answer = f"🔍 Response format error: {str(e)}"
        except Exception as e:
            answer = f"💥 Critical: {str(e)}"

    else:
        answer = f"❌ Unsupported provider: {server_type}"

    add_message(messages, "assistant", answer)
    return answer




def clear_screen() -> None:
    """Очистка экрана консоли"""
    console.clear()

def show_header() -> None:
    """Отображение заголовка программы"""
    clear_screen()
    title = Panel.fit(
        "🔍 [bold cyan]FILE SEARCHER[/]", 
        subtitle="by Python Powered",
        style="bold blue",
        width=MOBILE_WIDTH  # Ширина для панели
    )
    console.print(title)

def exit_program(message: str, code: int = 1) -> None:
    """Аварийное завершение программы"""
    console.print(f"\n[bold red]⛔ {message}[/]")
    sys.exit(code)

def get_input(prompt: str) -> str:
    """Получение ввода от пользователя с обработкой ошибок"""
    try:
        # Убраны неподдерживаемые параметры width и password
        return console.input(f"[bold green]▶ {prompt}: [/]").strip()
    except (EOFError, KeyboardInterrupt):
        exit_program("Operation cancelled")
    except Exception as e:
        exit_program(f"Input error: {str(e)}")

def show_error(message: str) -> None:
    """Отображение сообщения об ошибке"""
    console.print(f"[red]✖ ERROR: {message}[/]")

def highlight_text(text: str, search_text: str) -> str:
    """Подсветка найденного текста"""
    lower_text = text.lower()
    lower_search = search_text.lower()
    start = lower_text.find(lower_search)
    
    if start == -1:
        return text
    
    end = start + len(search_text)
    return (
        text[:start] +
        f"[{highlight_style}]{text[start:end]}[/]" +
        text[end:]
    )

def parse_json(data: dict, search_text: str, path: str = "") -> List:
    """Рекурсивный парсинг JSON-структур"""
    results = []
    search_lower = search_text.lower()
    
    if isinstance(data, dict):
        for key, value in data.items():
            new_path = f"{path}.{key}" if path else key
            results += parse_json(value, search_text, new_path)
    elif isinstance(data, list):
        for index, item in enumerate(data):
            new_path = f"{path}[{index}]"
            results += parse_json(item, search_text, new_path)
    else:
        str_value = str(data)
        if search_lower in str_value.lower():
            highlighted = highlight_text(str_value, search_text)
            results.append(("JSON", path, highlighted))
    
    return results

def process_file(file_path: str, search_term: str, find_first_only=True) -> Dict:
    """
    Обработка одного файла: сначала проверяет наличие плагина для расширения,
    если нет — использует стандартные методы поиска для txt, sql, xlsx, csv, json.
    Возвращает словарь {'path': file_path, 'matches': ...} или None.
    """

    ext = os.path.splitext(file_path)[1].lower()
    results = []

    # Обработка через плагин или стандартные обработчики
    if ext in file_handlers:
        try:
            plugin_info = file_handlers[ext]
            handler = plugin_info["handler"]
            import inspect
            if "find_first_only" in inspect.signature(handler).parameters:
                plugin_results = handler(file_path, search_term, find_first_only=find_first_only)
            else:
                plugin_results = handler(file_path, search_term)
            if plugin_results:
                return {'path': file_path, 'matches': plugin_results}
            else:
                return None
        except Exception as e:
            return {'path': file_path, 'matches': [("Error", "", f"[red]PLUGIN ERROR: {type(e).__name__}: {e}[/red]")]}
    # Стандартные форматы
    try:
        if ext in ('.txt', '.sql'):
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                for line_num, line in enumerate(f, 1):
                    if search_term.lower() in line.lower():
                        # Вывод всей строки
                        highlighted = highlight_text(line.strip(), search_term)
                        results.append(("Text", f"Line {line_num}", highlighted))
                        if find_first_only:
                            break

        elif ext == '.xlsx':
            from openpyxl import load_workbook
            wb = load_workbook(file_path, read_only=True)
            for sheet in wb:
                for row_num, row in enumerate(sheet.iter_rows(), 1):
                    if any(search_term.lower() in str(cell.value).lower() for cell in row if cell.value):
                        # Вывод всей строки
                        processed = [
                            highlight_text(str(cell.value), search_term) if cell.value and search_term.lower() in str(cell.value).lower() else str(cell.value)
                            for cell in row
                        ]
                        results.append(("Excel", f"{sheet.title} (Row {row_num})", " | ".join(processed)))
                        if find_first_only:
                            break
                if results and find_first_only:
                    break

        elif ext == '.csv':
            import csv
            with open(file_path, 'r', encoding='utf-8', errors='ignore') as f:
                reader = csv.reader(f)
                for row_num, row in enumerate(reader, 1):
                    if any(search_term.lower() in cell.lower() for cell in row):
                        processed = [highlight_text(cell, search_term) for cell in row]
                        results.append(("CSV", f"Row {row_num}", " | ".join(processed)))
                        if find_first_only:
                            break

        elif ext == '.json':
            import json
            def parse_json(data, search_text, path=""):
                results = []
                if isinstance(data, dict):
                    for key, value in data.items():
                        results += parse_json(value, search_text, f"{path}.{key}" if path else key)
                elif isinstance(data, list):
                    for index, item in enumerate(data):
                        results += parse_json(item, search_text, f"{path}[{index}]")
                else:
                    str_value = str(data)
                    if search_text.lower() in str_value.lower():
                        highlighted = highlight_text(str_value, search_text)
                        results.append(("JSON", path, highlighted))
                return results

            with open(file_path, 'r', encoding='utf-8') as f:
                data = json.load(f)
                results.extend(parse_json(data, search_term))
    except Exception as e:
        return {'path': file_path, 'matches': [("Error", "", f"[red]{type(e).__name__}: {e}[/red]")]}

    if results:
        return {'path': file_path, 'matches': results}
    else:
        return None
        
#    except Exception as e:
#        # Любая непредвиденная ошибка
#        return {'path': file_path, 'matches': [("Error", "", f"[red]{type(e).__name__}: {e}[/red]")]}


def display_results(results: List) -> None:
    """Отображение результатов в таблице"""
    if not results:
        console.print("[yellow]No matches found[/]", width=MOBILE_WIDTH)
        return

    table = Table(
        show_header=True,
        header_style="bold magenta",
        box=None,
        width=MOBILE_WIDTH
    )
    table.add_column("Type", width=8)
    table.add_column("Position", width=12)
    table.add_column("Content")

    for result in results:
        location_type, position, content = result
        table.add_row(
            Text(location_type, style="cyan"),
            Text(position, style="green"),
            Text.from_markup(content)
        )
    
    console.print(table)

def main_search(search_path: str, search_term: str, max_size_mb: float) -> Tuple[int, int, List]:
    """Оптимизированный поиск с фильтрацией по размеру"""
    total_files = 0
    found_count = 0
    all_results = []
    search_term = search_term.lower()
    
    # Предварительный сбор файлов с фильтрацией
    file_list = []
    for root, _, files in os.walk(search_path):
        for file in files:
            file_path = os.path.join(root, file)
            ext = os.path.splitext(file)[1].lower()
            file_size = os.path.getsize(file_path) / (1024 * 1024)  # Размер в MB
            
            if ext in SUPPORTED_EXTENSIONS and file_size <= max_size_mb:
                file_list.append(file_path)
    
    total_files = len(file_list)
    
    # Обработка файлов с ThreadPool
    with ThreadPoolExecutor(max_workers=MAX_WORKERS) as executor:
        futures = []
        for file_path in file_list:
            futures.append(executor.submit(process_file, file_path, search_term))
        
        for future in futures:
            result = future.result()
            if result:
                found_count += 1
                all_results.append(result)
                console.print(Panel.fit(
                    f"[bold cyan]{result['path']}[/]",
                    style="blue"
                ))
                display_results(result['matches'])
    
    return total_files, found_count, all_results

def save_results(data: List, search_text: str, total: int, found: int) -> None:
    """Сохранение результатов в файл"""
    clean_text = "".join([c if c.isalnum() else "_" for c in search_text])
    filename = os.path.join(REPORTS_DIR,f"search_{clean_text}_{datetime.now():%Y%m%d_%H%M%S}.txt")
    
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write("-- SEARCH REPORT\n")
            f.write(f"-- Search term: '{search_text}'\n")
            f.write(f"-- Date: {datetime.now():%Y-%m-%d %H:%M:%S}\n\n")
            f.write("CREATE TABLE IF NOT EXISTS search_results (\n")
            f.write("    id INTEGER PRIMARY KEY,\n")
            f.write("    file_path TEXT,\n")
            f.write("    result_type TEXT,\n")
            f.write("    position TEXT,\n")
            f.write("    content TEXT\n);\n\n")
            
            for entry in data:
                for match in entry['matches']:
                    content = match[2].replace("'", "''")
                    f.write(f"INSERT INTO search_results (file_path, result_type, position, content)\n")
                    f.write(f"VALUES ('{entry['path']}', '{match[0]}', '{match[1]}', '{content}');\n")
        
        console.print(f"\n[bold green]✓ Report saved to: [u]{filename}[/][/]")
    except Exception as e:
        show_error(f"Failed to save report: {str(e)}")

def save_to_sql(data: List, search_text: str) -> str:
    """Сохранение результатов в SQL файл"""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    clean_text = "".join([c if c.isalnum() else "_" for c in search_text])
    filename = os.path.join(REPORTS_DIR, f"search_{clean_text}_{datetime.now():%Y%m%d_%H%M%S}.sql")
    
    try:
        with open(filename, 'w', encoding='utf-8') as f:
            f.write("-- SEARCH REPORT\n")
            f.write(f"-- Search term: '{search_text}'\n")
            f.write(f"-- Date: {datetime.now():%Y-%m-%d %H:%M:%S}\n\n")
            f.write("CREATE TABLE IF NOT EXISTS search_results (\n")
            f.write("    id INTEGER PRIMARY KEY,\n")
            f.write("    file_path TEXT,\n")
            f.write("    result_type TEXT,\n")
            f.write("    position TEXT,\n")
            f.write("    content TEXT\n);\n\n")
            
            for entry in data:
                for match in entry['matches']:
                    content = match[2].replace("'", "''")
                    f.write(f"INSERT INTO search_results (file_path, result_type, position, content)\n")
                    f.write(f"VALUES ('{entry['path']}', '{match[0]}', '{match[1]}', '{content}');\n")
        
        return filename
    except Exception as e:
        show_error(f"Error saving SQL: {str(e)}")
        return ""

def show_sql_reports():
    """Показать список доступных SQL-отчетов"""
    if not os.path.exists(REPORTS_DIR):
        console.print("[yellow]No reports found[/]")
        return

    files = [f for f in os.listdir(REPORTS_DIR) if f.endswith('.sql')]
    if not files:
        console.print("[yellow]No SQL reports found[/]")
        return

    console.print("\n[bold]Available reports:[/]")
    for i, file in enumerate(files, 1):
        console.print(f"{i}. {file}")

    choice = get_input("Select report number (0 to cancel)")
    if choice.isdigit() and 0 < int(choice) <= len(files):
        selected = os.path.join(REPORTS_DIR, files[int(choice)-1])
        with open(selected, 'r', encoding='utf-8') as f:
            console.print(f"\n[bold]Contents of {files[int(choice)-1]}:[/]\n")
            console.print(f.read())
    else:
        console.print("[yellow]Operation cancelled[/]")

def save_to_json(data: List, search_text: str) -> str:
    """Сохранение результатов в JSON файл"""
    os.makedirs(REPORTS_DIR, exist_ok=True)
    clean_text = "".join([c if c.isalnum() else "_" for c in search_text])
    filename = os.path.join(REPORTS_DIR, f"search_{clean_text}_{datetime.now():%Y%m%d_%H%M%S}.json")
    
    try:
        report_data = {
            "search_term": search_text,
            "timestamp": datetime.now().isoformat(),
            "results": [
                {
                    "file": entry["path"],
                    "matches": [
                        {
                            "type": match[0],
                            "position": match[1],
                            "content": match[2].replace('[', '').replace(']', '')
                        } 
                        for match in entry["matches"]
                    ]
                } 
                for entry in data
            ]
        }
        
        with open(filename, 'w', encoding='utf-8') as f:
            json.dump(report_data, f, indent=2, ensure_ascii=False)
        
        return filename
    except Exception as e:
        show_error(f"Error saving JSON: {str(e)}")
        return ""

def show_json_reports():
    """Показать список доступных json-отчетов"""
    if not os.path.exists(REPORTS_DIR):
        console.print("[yellow]No reports found[/]")
        return

    files = [f for f in os.listdir(REPORTS_DIR) if f.endswith('.json')]
    if not files:
        console.print("[yellow]No json reports found[/]")
        return

    console.print("\n[bold]Available reports:[/]")
    for i, file in enumerate(files, 1):
        console.print(f"{i}. {file}")

    choice = get_input("Select report number (0 to cancel)")
    if choice.isdigit() and 0 < int(choice) <= len(files):
        selected = os.path.join(REPORTS_DIR, files[int(choice)-1])
        with open(selected, 'r', encoding='utf-8') as f:
            console.print(f"\n[bold]Contents of {files[int(choice)-1]}:[/]\n")
            console.print(f.read())
    else:
        console.print("[yellow]Operation cancelled[/]")

def show_txt_reports():
    """Показать список доступных txt-отчетов"""
    if not os.path.exists(REPORTS_DIR):
        console.print("[yellow]No reports found[/]")
        return

    files = [f for f in os.listdir(REPORTS_DIR) if f.endswith('.json')]
    if not files:
        console.print("[yellow]No txt reports found[/]")
        return

    console.print("\n[bold]Available reports:[/]")
    for i, file in enumerate(files, 1):
        console.print(f"{i}. {file}")

    choice = get_input("Select report number (0 to cancel)")
    if choice.isdigit() and 0 < int(choice) <= len(files):
        selected = os.path.join(REPORTS_DIR, files[int(choice)-1])
        with open(selected, 'r', encoding='utf-8') as f:
            console.print(f"\n[bold]Contents of {files[int(choice)-1]}:[/]\n")
            console.print(f.read())
    else:
        console.print("[yellow]Operation cancelled[/]")

def main_menu():
    """Главное меню программы"""
    load_plugins()  # обязательно!    
    while True:
        show_header()
        console.print(Panel.fit("1. New search\n2. View reports\n3. Show plugins\n4. Ask ai\n5. Cleaner\n6. Exit", 
                              style="cyan",
                              width=MOBILE_WIDTH))
        
        choice = get_input("Select option")
        if choice == '1':
            return 'search'
        elif choice == '2':
            show_sql_reports()
            show_json_reports()
            show_txt_reports()
            get_input("\nPress Enter to return to menu...")
        elif choice == "5":
            # Запуск Cleaner
            if "cleaner" in file_handlers:
                file_handlers["cleaner"]["handler"]()
            else:
                print("Cleaner not available!")            
        elif choice == '6':
            exit_program("Goodbye!")
        else:
            show_error("Invalid choice")
            clear_screen()

def main() -> None:
    """Главная функция программы"""
    load_plugins()  # обязательно!
    while True:
        # Основной цикл меню
        while True:
            show_header()
            console.print(Panel.fit(
                "1. New search\n2. View reports\n3. Show plugins\n4. Ask ai\n5. Cleaner\n6. Exit",
                style="cyan",
                width=MOBILE_WIDTH
            ))
            
            choice = get_input("Select option")
            
            if choice == '1':
                break  # Выход из меню для начала поиска
            elif choice == '2':
                show_sql_reports()
                show_json_reports()
                show_txt_reports()
                get_input("\nPress Enter to return to menu...")
            elif choice == '3':
                show_plugins_info()
            elif choice == '4':
                SERVER_TYPE = get_input("Choose a server (google/huggingface): ").lower()
                if SERVER_TYPE not in ('google', 'huggingface'):
                    show_error("Invalid server choice")
                    continue
                    
                # Ввод вопроса
                prompt = get_input("Enter your question: ")    
                
                # Отправка запроса
                response = ask_ai(prompt, SERVER_TYPE)
                
                # Вывод результата
                console.clear()
                console.print(Panel.fit(response, style="blue", title="AI Response"))
                get_input("\nPress Enter to return to menu...")
            
            elif choice == "5":
                # Запуск Cleaner
                if "cleaner" in file_handlers:
                    file_handlers["cleaner"]["handler"]()
                else:
                    print("Cleaner not available!")    
            elif choice == '6':
                exit_program("Goodbye!")
            else:
                show_error("Invalid choice")
                clear_screen()

        # Блок поиска
        search_path = ""
        while True:
            search_path = get_input("Search directory")
            if not search_path:
                show_error("Path cannot be empty")
                continue
            if not os.path.exists(search_path):
                show_error("Path does not exist")
                continue
            if not os.path.isdir(search_path):
                show_error("Must be a directory")
                continue
            break

        # Поисковый запрос
        search_term = ""
        while True:
            search_term = get_input("Search term")
            if not search_term:
                show_error("Search term required")
                continue
            if len(search_term) < 2:
                show_error("Minimum 2 characters")
                continue
            break

        console.clear()
        show_header()
        
        try:
            # Выполнение поиска
            console.print(f"[bold]Searching for: [u green]'{search_term}'[/][/]")
            total, found, results = main_search(search_path, search_term.lower(), max_size)

            # Вывод результатов
            console.print(f"\n[bold cyan]■ Search Complete ■[/]")
            console.print(f"│ Scanned: [green]{total} files[/]")
            console.print(f"│ Matches: [green]{found} files[/]")

            # Блок сохранения результатов
            if found > 0:
                save_choice = get_input("Save results? (Y/n)").lower()
                if save_choice in ('', 'y', 'yes'):
                    console.print("\n1. TXT\n2. SQL\n3. JSON")
                    format_choice = get_input("Select format")
                    
                    if format_choice == '1':
                        save_results(results, search_term, total, found)
                        console.print(f"\n[bold green]✓ TXT report saved![/]")
                    elif format_choice == '2':
                        filename = save_to_sql(results, search_term)
                        if filename:
                            console.print(f"\n[bold green]✓ SQL report saved to: [u]{filename}[/][/]")
                    elif format_choice == '3':
                        filename = save_to_json(results, search_term)
                        if filename:
                            console.print(f"\n[bold green]✓ JSON report saved to: [u]{filename}[/][/]")
                    else:
                        console.print("[yellow]Export cancelled[/]")

            # Возврат в меню
            get_input("\nPress Enter to return to menu...")

        except KeyboardInterrupt:
            exit_program("Search cancelled by user")
        except Exception as e:
            show_error(f"Search error: {str(e)}")
            get_input("\nPress Enter to return to menu...")

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        exit_program(f"Critical error: {str(e)}")